import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io
import base64

# Configure the page
st.set_page_config(
    page_title="Processeur Excel V2",
    layout="wide"
)

# Add logo to top corner
def get_base64_encoded_image(image_path):
    with open(image_path, "rb") as img_file:
        return base64.b64encode(img_file.read()).decode()

# Create two columns - one for logo, one for title
col1, col2 = st.columns([1, 4])

with col1:
    try:
        logo_base64 = get_base64_encoded_image("logo.png")
        st.markdown(
            f'<img src="data:image/png;base64,{logo_base64}" width="100">',
            unsafe_allow_html=True
        )
    except:
        st.error("Fichier logo non trouvé. Veuillez vous assurer que 'logo.png' est dans le même répertoire que le script.")

st.title("Remplissage de la fiche de lancement")
#st.write("Téléchargez vos fichiers Excel pour les traiter")

# Create columns for file uploaders
col1, col2 = st.columns(2)

# File uploaders in columns
with col1:
    st.markdown("**Fichier Stock**")
    st.markdown("Sélectionnez votre fichier de stock")
    stock_file = st.file_uploader("", type=['xlsx'], key="stock_uploader")

with col2:
    st.markdown("**Fichier Lancement**")
    st.markdown("Sélectionnez votre fiche de lancement")
    launch_file = st.file_uploader("", type=['xlsx'], key="launch_uploader")

if stock_file is not None and launch_file is not None:
    try:
        # Load stock data
        stock_df = pd.read_excel(stock_file, sheet_name="Feuil1")
        stock_df.set_index("RÉF", inplace=True)

        # Load workbook for reading values
        wb_read = load_workbook(launch_file, data_only=True)
        ws_read = wb_read["CHEF PRODUIT"]

        # Load workbook for preserving formulas
        wb = load_workbook(launch_file)
        ws = wb["CHEF PRODUIT"]

        # Define the start row for references and the column mappings
        start_row = 29
        ref_col = "G"
        color_col = "D"  # Column for color
        laize_col = "H"  # Column for laize

        # Mapping: column in Excel sheet -> column in stock_df
        column_mappings = {
            "H": "Laize",
            "I": "Composition",
            "J": "P/M²",
            "L": "FRNS",
            "R": "Prix Dh",
        }

        # Progress bar
        progress_bar = st.progress(0)
        status_text = st.empty()
        debug_text = st.empty()

        # Loop through rows starting from G29 down
        row = start_row
        total_rows = 0
        while True:
            ref_cell = f"{ref_col}{row}"
            reference = ws_read[ref_cell].value

            if reference is None:
                break  # Stop when we hit an empty row
            total_rows += 1
            row += 1

        row = start_row
        processed_rows = 0
        while True:
            ref_cell = f"{ref_col}{row}"
            reference = ws_read[ref_cell].value

            if reference is None:
                break  # Stop when we hit an empty row

            if reference in stock_df.index:
                row_data = stock_df.loc[reference]
                
                # Check if quantity is zero
                if int(row_data.get("Qte Phys Reel", 0)) == 0:
                    debug_text.text(f"⚠️ Référence '{reference}' a une quantité nulle. Recherche d'alternatives...")
                    # Get current color and laize
                    current_color = row_data.get("COL", "")
                    current_laize = row_data.get("Laize", "")
                    
                    debug_text.text(f"Recherche d'alternatives avec:\nCouleur: '{current_color}'\nLaize: '{current_laize}'")
                    
                    # Find alternative reference with same color and laize
                    mask = (
                        stock_df["COL"].fillna("").astype(str).str.strip().str.lower() == str(current_color).strip().lower()
                    ) & (
                        stock_df["Laize"].fillna("").astype(str).str.strip().str.lower() == str(current_laize).strip().lower()
                    ) & (
                        pd.to_numeric(stock_df["Qte Phys Reel"].fillna(0), errors='coerce') > 0
                    )
                    
                    alternative_refs = stock_df[mask]
                    
                    if len(alternative_refs) > 0:
                        # Use the first alternative reference found
                        reference = alternative_refs.index[0]
                        row_data = stock_df.loc[reference]
                        status_text.text(f"⚠️ Référence originale '{reference}' a une quantité nulle. Utilisation de la référence alternative '{reference}' avec la même couleur et laize.")
                    else:
                        status_text.text(f"❌ Référence '{reference}' a une quantité nulle et aucune alternative trouvée avec la même couleur et laize.")
                        row += 1
                        processed_rows += 1
                        progress_bar.progress(processed_rows / total_rows)
                        continue

                # Process the data
                for col_letter, stock_column in column_mappings.items():
                    target_cell = f"{col_letter}{row}"
                    value = row_data.get(stock_column, "")
                    # Remove 'dh' from price values regardless of position
                    if col_letter == "R" and isinstance(value, str):
                        value = value.replace("dh", "").strip()
                    ws[target_cell] = value
                status_text.text(f"✅ Référence '{reference}' trouvée dans le fichier stock (ligne {row})")
            else:
                status_text.text(f"❌ Référence '{reference}' non trouvée dans le fichier stock (ligne {row})")

            row += 1
            processed_rows += 1
            progress_bar.progress(processed_rows / total_rows)

        # Save the updated workbook to a BytesIO object
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create download button
        st.download_button(
            label="Télécharger le fichier traité",
            data=output,
            file_name="fiche_lancement_complete.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.success("Traitement terminé avec succès!")

    except Exception as e:
        st.error(f"Une erreur s'est produite: {str(e)}") 